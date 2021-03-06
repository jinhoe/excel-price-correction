import openpyxl as xl                           # Import module
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)             # Load file
    sheet = wb['Sheet1']                        # Define sheet
    # cell = sheet['a1']                        # Example define cell row 1, column 1
    # cell = sheet.cell(1, 1)                   # Example define cell row 1, column 1

    for row in range(2, sheet.max_row + 1):     # Parse from row 2 to 4
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    corrected_price_header_cell = sheet.cell(1, 4)
    corrected_price_header_cell.value = 'corrected_price'

    values = Reference(sheet,                   # Select a range of data row 2 to 4(max), column 4
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename[:-5] + '_corrected.xlsx')

process_workbook('transactions.xlsx')
